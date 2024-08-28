from ...models import Wallet, Currency, Transaction,Permission, ResourceAssignment, Resource, ConsumingResource, Task\
    , PayOff, TransactionAdditionalReceipt, OPArea, OPProject
from ...Serializers.wallet_serializer import TransactionSerializer
from ...views.time.time import AddTimeAndReport
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Q,Window,F, Value, IntegerField, CharField, DecimalField, Sum
from django.db.models.functions.window import FirstValue
from django.core.exceptions import PermissionDenied
from rest_framework.renderers import JSONRenderer
from ...utilities.date_tools import  ConvertToSolarDate
import json
import decimal
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from zipfile import ZipFile
from io import BytesIO
from django.conf import settings
import os
import string
from jdatetime import datetime as jdt

# wallet management main page render
@login_required(login_url='user:login') #redirect when user is not logged in
def index(request):
    if not request.user.employee.WalletPageAccess :
        raise PermissionDenied
    request.session["activated_menu"]="wallet_manage"
    context={}

    _archive = False
    if request.GET and request.GET['archive'] == 'True':
        _archive = True
    context["switch_archive"] = _archive

    tasks = Task.objects.filter(
        Q(user_assignee=request.user)| Q(group_assignee__head=request.user)| Q(public = True)).exclude(cancelled=True).exclude(progress__lt=1).exclude(progress__gt=99)
    context['task_list'] = tasks

    context['has_transaction_access'] = False
    if request.user.pk in Permission.objects.filter(codename = 'add_transaction')[0].user_set.all().values_list('pk', flat = True):
        context['has_transaction_access'] = True
    
    context['currency']=Currency.objects.all()
    context['wallets']=Wallet.objects.filter(active = True, archived = _archive)
    context['master_wallets']=Wallet.objects.filter(active = True, archived = _archive, master = True)

    
    if(request.user.employee.organization_group.manager==request.user):
        context["isManager"]=True
    else:
        context["isManager"]=False


    # find all resources needed for this page
    if len(request.user.locumtenens_organization_groups.all())>0 and request.user.locumtenens_organization_groups.first().locumtenens_active:
        as_user = request.user.locumtenens_organization_groups.first().manager
    else:
        as_user = request.user

    # useful variables defined to increase performance

    as_user_employee_GetAllChildrenUserId = as_user.employee.GetAllChildrenUserId
    as_user_assign_list = ResourceAssignment.objects.filter(assignee=as_user,deleted=None).values_list('resource__pk',flat=True)
    as_user_employee_TaskAssignedResources = as_user.employee.TaskAssignedResources
    as_user_assign_list_child = ResourceAssignment.objects.filter(assignee__id__in=as_user_employee_GetAllChildrenUserId,deleted=None).values_list('resource__pk',flat=True)
    total_assign_pk = set(as_user_assign_list) | set(as_user_employee_TaskAssignedResources) | set(as_user_assign_list_child)

    _resources_no_expire_id_list= ConsumingResource.objects.filter(expiration__gte=datetime.date.today()).values_list('resource__pk',flat=True)
    _resources_no_expire_id_list = set(_resources_no_expire_id_list)
    if None in _resources_no_expire_id_list:
        _resources_no_expire_id_list.remove(None)

    _software=Resource.objects.filter(
        Q(owner=as_user) |Q(owner__id__in=as_user_employee_GetAllChildrenUserId)|Q(locumtenens=as_user)|Q(locumtenens__id__in=as_user_employee_GetAllChildrenUserId)|
        Q(pk__in=total_assign_pk) 
        ,resource_type__category=3)
         
    _consuming=Resource.objects.filter(pk__in=_resources_no_expire_id_list).filter(
    Q(owner=as_user) |Q(owner__id__in=as_user_employee_GetAllChildrenUserId)|Q(locumtenens=as_user)|Q(locumtenens__id__in=as_user_employee_GetAllChildrenUserId)|
    Q(pk__in=total_assign_pk)
    ,resource_type__category=1).annotate(consuming_expiration=Window(
        expression=FirstValue('consuming_resources__expiration'),
        partition_by=[F('consuming_resources__resource__id'),],
        order_by=F('consuming_resources__id').desc()
    )).distinct()



    _consuming=_consuming.exclude(deleted=True)


    context["consuming_resources"]=_consuming.only('id','name')

    context["software_resources"]=_software.only('id','name')

    return render(request, 'wallet/index.html', {'context':context})

# update currencies Rial ration from wallet management main page
@login_required(login_url='user:login') #redirect when user is not logged in
@csrf_exempt
def currency(request):
    if request.method == 'POST':
        if request.user.pk in Permission.objects.filter(codename = 'add_transaction')[0].user_set.all().values_list('pk', flat = True):
            new_values = json.loads(request.body.decode('utf-8'))

            try:
                for cur_value in new_values:
                    cur_id = int(cur_value['id'])
                    cur_val = decimal.Decimal(cur_value['value'])
                    if Currency.objects.filter(pk=cur_id).exists() :
                        cur = Currency.objects.get(pk=cur_id)
                        cur.rialratio = cur_val
                        cur.save()
                        

                return JsonResponse({'Result':'OK'})
            except:
                return JsonResponse({'Result':'Fail'})
        else:
            return JsonResponse({'Result':'Fail'})

# register new transactions in datebase
@login_required(login_url='user:login') #redirect when user is not logged in
@csrf_exempt
@transaction.atomic
def transaction(request, ttype):
    if request.method == 'POST' and request.user.pk in Permission.objects.filter(codename = 'add_transaction')[0].user_set.all().values_list('pk', flat = True):
        try:
            if ttype == 1 :
                if int(request.POST['transfer_source_wallet']) == int(request.POST['transfer_destination_wallet']):
                    return JsonResponse({'message':"مبدا و مقصد نمی تواند یکسان باشد"}, status = 424)

                # if Wallet.objects.get(pk = int(request.POST['transfer_source_wallet'])).currency != Wallet.objects.get(pk = int(request.POST['transfer_destination_wallet'])).currency :
                #     return JsonResponse({'message':"مبدا و مقصد نمی تواند واحد پول متفاوت داشته باشد"}, status = 424)

                source_transaction = Transaction()
                source_transaction.wallet = Wallet.objects.get(pk = int(request.POST['transfer_source_wallet']))
                source_transaction.payoff_wallet = Wallet.objects.get(pk = int(request.POST['transfer_wallet_payoff']))
                source_transaction.amount = decimal.Decimal(request.POST['transfer_source_amount'])
                if 'transfer_source_amount_dollar' in request.POST and request.POST['transfer_source_amount_dollar']:
                    source_transaction.amount_dollar = decimal.Decimal(request.POST['transfer_source_amount_dollar'])
                source_transaction.incordec = 1
                source_transaction.fee = decimal.Decimal(request.POST['transfer_source_fee'])
                if 'transfer_source_fee_dollar' in request.POST and request.POST['transfer_source_fee_dollar'] :
                    source_transaction.fee_dollar = decimal.Decimal(request.POST['transfer_source_fee_dollar'])
                source_transaction.wallet_balance_after_d = source_transaction.wallet.current_balance - (source_transaction.fee + source_transaction.amount)
                source_transaction.time = datetime.datetime.fromisoformat(request.POST['transfer_source_datetime'])
                source_transaction.receipt_file = request.FILES['transfer_source_receipt']
                source_transaction.receipt_file_name = request.FILES['transfer_source_receipt'].name
                source_transaction.creator = request.user
                if len(request.POST['transfer_source_comment']) > 0 :
                    source_transaction.comment = request.POST['transfer_source_comment']
                
                source_transaction.save()

                if len(request.FILES.getlist('transfer_source_receipt_additional')) > 0:
                    for add_file in request.FILES.getlist('transfer_source_receipt_additional'):
                        additional_file = TransactionAdditionalReceipt()
                        additional_file.receipt_file = add_file
                        additional_file.receipt_file_name = add_file.name
                        additional_file.transaction = source_transaction
                        additional_file.creator = request.user
                        additional_file.save()

                source_transaction.wallet.current_balance_d = source_transaction.wallet_balance_after
                source_transaction.wallet.save()

                destination_transaction = Transaction()
                destination_transaction.wallet = Wallet.objects.get(pk = int(request.POST['transfer_destination_wallet']))
                destination_transaction.payoff_wallet = Wallet.objects.get(pk = int(request.POST['transfer_wallet_payoff']))
                destination_transaction.amount = decimal.Decimal(request.POST['transfer_destination_amount'])
                if 'transfer_source_amount_dollar' in request.POST and request.POST['transfer_source_amount_dollar'] :
                    destination_transaction.amount_dollar = decimal.Decimal(request.POST['transfer_source_amount_dollar'])
                destination_transaction.incordec = 0
                destination_transaction.fee = 0
                destination_transaction.wallet_balance_after_d = destination_transaction.wallet.current_balance + destination_transaction.amount
                destination_transaction.time = datetime.datetime.fromisoformat(request.POST['transfer_source_datetime'])
                if request.FILES.get('transfer_destination_receipt'):
                    destination_transaction.receipt_file = request.FILES['transfer_destination_receipt']
                    destination_transaction.receipt_file_name = request.FILES['transfer_destination_receipt'].name
                else:
                    destination_transaction.receipt_file = request.FILES['transfer_source_receipt']
                    destination_transaction.receipt_file_name = request.FILES['transfer_source_receipt'].name
                destination_transaction.creator = request.user
                if len(request.POST['transfer_destination_comment']) > 0 :
                    destination_transaction.comment = request.POST['transfer_destination_comment']
                destination_transaction.source_transaction = source_transaction

                destination_transaction.save()

                destination_transaction.wallet.current_balance_d = destination_transaction.wallet_balance_after
                destination_transaction.wallet.save()

                if "transfer_auto_report" in request.POST:
                    if "transfer_report_task" in request.POST and "transfer_report_start" in request.POST and "transfer_report_end" in request.POST:
                        _task = Task.objects.get(pk = int(request.POST['transfer_report_task']))
                        if _task.user_assignee == request.user or _task.group_assignee.head == request.user:
                            report_text = "تبادل پول از کیف پول " + source_transaction.wallet.name + " به کیف پول " + destination_transaction.wallet.name + " انجام شد "
                            start = datetime.time.fromisoformat(request.POST['transfer_report_start'])
                            end = datetime.time.fromisoformat(request.POST['transfer_report_end'])
                            today = datetime.datetime.today()
                            start_datetime = datetime.datetime(year= today.year, month= today.month, day=today.day, hour=start.hour, minute= start.minute)
                            end_datetime = datetime.datetime(year= today.year, month= today.month, day=today.day, hour=end.hour, minute= end.minute)
                            return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت و " + AddTimeAndReport(_task.pk,start_datetime,end_datetime,report_text)})
                        else:
                            return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت شد ولی کار انتخابی برای گزارش مربوط به کاربر نیست"})
                    else:
                        return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت شد ولی گزارش آن به علت نقص اطلاعات ثبت نشد"})

                return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت شد"})
        
            if ttype == 2 :
                pay_transaction = Transaction()
                if 'payment_title' in request.POST and len(request.POST['payment_title']) > 0:
                    pay_transaction.title = request.POST['payment_title']
                pay_transaction.wallet = Wallet.objects.get(pk = int(request.POST['payment_wallet']))
                pay_transaction.payoff_wallet = Wallet.objects.get(pk = int(request.POST['payment_wallet_payoff']))
                pay_transaction.amount = decimal.Decimal(request.POST['payment_amount'])
                if 'payment_amount_dollar' in request.POST and request.POST['payment_amount_dollar'] :
                    pay_transaction.amount_dollar = decimal.Decimal(request.POST['payment_amount_dollar'])
                pay_transaction.incordec = 1
                pay_transaction.fee = decimal.Decimal(request.POST['payment_fee'])
                if 'payment_fee_dollar' in request.POST and request.POST['payment_fee_dollar'] :
                    pay_transaction.fee_dollar = decimal.Decimal(request.POST['payment_fee_dollar'])
                pay_transaction.wallet_balance_after_d = pay_transaction.wallet.current_balance - (pay_transaction.amount + pay_transaction.fee)
                pay_transaction.time = datetime.datetime.fromisoformat(request.POST['payment_datetime'])
                pay_transaction.receipt_file = request.FILES['payment_receipt']
                pay_transaction.receipt_file_name = request.FILES['payment_receipt'].name
                pay_transaction.creator = request.user
                if len(request.POST['payment_comment']) > 0 :
                    pay_transaction.comment = request.POST['payment_comment']
                
                pay_transaction.dest_resource = Resource.objects.get(pk = int(request.POST['payment_resource']))
                pay_transaction.save()

                if len(request.FILES.getlist('payment_receipt_additional')) > 0:
                    for add_file in request.FILES.getlist('payment_receipt_additional'):
                        additional_file = TransactionAdditionalReceipt()
                        additional_file.receipt_file = add_file
                        additional_file.receipt_file_name = add_file.name
                        additional_file.transaction = pay_transaction
                        additional_file.creator = request.user
                        additional_file.save()

                pay_transaction.wallet.current_balance_d = pay_transaction.wallet_balance_after
                pay_transaction.wallet.save()

                if "payment_auto_report" in request.POST:
                    if "payment_report_task" in request.POST and "payment_report_start" in request.POST and "payment_report_end" in request.POST:
                        _task = Task.objects.get(pk = int(request.POST['payment_report_task']))
                        if _task.user_assignee == request.user or _task.group_assignee.head == request.user:
                            report_text = "خرید از کیف پول " + pay_transaction.wallet.name + " برای منبع " + pay_transaction.dest_resource.name + " انجام شد "
                            start = datetime.time.fromisoformat(request.POST['payment_report_start'])
                            end = datetime.time.fromisoformat(request.POST['payment_report_end'])
                            today = datetime.datetime.today()
                            start_datetime = datetime.datetime(year= today.year, month= today.month, day=today.day, hour=start.hour, minute= start.minute)
                            end_datetime = datetime.datetime(year= today.year, month= today.month, day=today.day, hour=end.hour, minute= end.minute)
                            return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت و " + AddTimeAndReport(_task.pk,start_datetime,end_datetime,report_text)})
                        else:
                            return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت شد ولی کار انتخابی برای گزارش مربوط به کاربر نیست"})
                    else:
                        return JsonResponse({'message':"تبادل انجام شده با موفقیت ثبت شد ولی گزارش آن به علت نقص اطلاعات ثبت نشد"})


                return JsonResponse({'message':"خرید انجام شده با موفقیت ثبت شد"})

            if ttype == 3 :
                

                # if Wallet.objects.get(pk = int(request.POST['transfer_source_wallet'])).currency != Wallet.objects.get(pk = int(request.POST['transfer_destination_wallet'])).currency :
                #     return JsonResponse({'message':"مبدا و مقصد نمی تواند واحد پول متفاوت داشته باشد"}, status = 424)

                transaction = Transaction.objects.get(pk = int(request.POST['transfer_edit_id']))
                if transaction.source_transaction :
                    rel_trans = transaction.source_transaction
                else :
                    rel_trans = transaction.dest_transaction
                if int(request.POST['transfer_edit_wallet']) == rel_trans.wallet.id:
                    return JsonResponse({'message':"مبدا و مقصد نمی تواند یکسان باشد"}, status = 424)
                transaction.wallet = Wallet.objects.get(pk = int(request.POST['transfer_edit_wallet']))
                transaction.payoff_wallet = Wallet.objects.get(pk = int(request.POST['transfer_edit_wallet_payoff']))
                rel_trans.payoff_wallet = Wallet.objects.get(pk = int(request.POST['transfer_edit_wallet_payoff']))
                transaction.amount = decimal.Decimal(request.POST['transfer_edit_amount'])
                if 'transfer_edit_amount_dollar' in request.POST and request.POST['transfer_edit_amount_dollar']:
                    transaction.amount_dollar = decimal.Decimal(request.POST['transfer_edit_amount_dollar'])

                transaction.fee = decimal.Decimal(request.POST['transfer_edit_fee'])
                if 'transfer_edit_fee_dollar' in request.POST and request.POST['transfer_edit_fee_dollar']:
                    transaction.fee_dollar = decimal.Decimal(request.POST['transfer_edit_fee_dollar'])
                transaction.wallet_balance_after_d = transaction.wallet.current_balance - (transaction.fee + transaction.amount)
                transaction.time = datetime.datetime.fromisoformat(request.POST['transfer_edit_datetime'])
                if 'transfer_edit_receipt' in request.FILES :
                    transaction.receipt_file = request.FILES['transfer_edit_receipt']
                    transaction.receipt_file_name = request.FILES['transfer_edit_receipt'].name
                
                if len(request.POST['transfer_edit_comment']) > 0 :
                    transaction.comment = request.POST['transfer_edit_comment']
                else:
                    transaction.comment = None


                transaction.save()
                rel_trans.save()

                transaction.wallet.current_balance_d = transaction.wallet_balance_after
                transaction.wallet.save()

                return JsonResponse({'message':"تغییر انجام شده با موفقیت ثبت شد"})
        
            if ttype == 4 :
                pay_transaction = Transaction.objects.get(pk = int(request.POST['payment_edit_id']))
                if len(request.POST['payment_edit_title']) > 0 :
                    pay_transaction.title = request.POST['payment_edit_title']
                else :
                    pay_transaction.title = None
                pay_transaction.wallet = Wallet.objects.get(pk = int(request.POST['payment_edit_wallet']))
                pay_transaction.payoff_wallet = Wallet.objects.get(pk = int(request.POST['payment_edit_wallet_payoff']))
                pay_transaction.amount = decimal.Decimal(request.POST['payment_edit_amount'])
                if 'payment_edit_amount_dollar' in request.POST and request.POST['payment_edit_amount_dollar']:
                    pay_transaction.amount_dollar = decimal.Decimal(request.POST['payment_edit_amount_dollar'])
                pay_transaction.fee = decimal.Decimal(request.POST['payment_edit_fee'])
                if 'payment_edit_fee_dollar' in request.POST and request.POST['payment_edit_fee_dollar']:
                    pay_transaction.fee_dollar = decimal.Decimal(request.POST['payment_edit_fee_dollar'])
                pay_transaction.wallet_balance_after_d = pay_transaction.wallet.current_balance - (pay_transaction.amount + pay_transaction.fee)
                pay_transaction.time = datetime.datetime.fromisoformat(request.POST['payment_edit_datetime'])
                if 'payment_edit_receipt' in request.FILES :
                    pay_transaction.receipt_file = request.FILES['payment_edit_receipt']
                    pay_transaction.receipt_file_name = request.FILES['payment_edit_receipt'].name
                pay_transaction.creator = request.user
                if len(request.POST['payment_edit_comment']) > 0 :
                    pay_transaction.comment = request.POST['payment_edit_comment']
                else :
                    pay_transaction.comment = None
                
                pay_transaction.dest_resource = Resource.objects.get(pk = int(request.POST['payment_edit_resource']))
                pay_transaction.save()

                pay_transaction.wallet.current_balance_d = pay_transaction.wallet_balance_after
                pay_transaction.wallet.save()

                return JsonResponse({'message':"تغییر انجام شده با موفقیت ثبت شد"})
        except:
            return JsonResponse({'message':"خطا در ثبت تراکنش"}, status = 424)
    else:
        return JsonResponse({'message':"تراکنش غیر مجاز"}, status = 424)
    

@login_required(login_url='user:login') #redirect when user is not logged in
@csrf_exempt
def transaction_list(request, wallet_id):
    if request.user.employee.WalletPageAccess :
        data = {}
        transactions = TransactionSerializer(Transaction.objects.filter(wallet= wallet_id), many = True)
        data['transactions'] = JSONRenderer().render(transactions.data).decode("utf-8")

        return JsonResponse(data)
    else:
        raise PermissionDenied
    

@login_required(login_url='user:login')
def archive_wallet(request,wallet_id):
    if request.user.employee.WalletPageAccess :
        try:
            wallet = Wallet.objects.get(pk = wallet_id)
            if wallet.current_balance == 0:
                wallet.archived = True
                wallet.save()
                return JsonResponse({'message':"آرشیو کردن کیف پول انجام شد"}, status = 200)
            else:
                return JsonResponse({'message':"آرشیو کردن کیف پول دارای موجودی ممکن نیست"}, status = 424)
        except:
            return JsonResponse({'message':"آرشیو کردن کیف پول با شکست مواجه شد"}, status = 424)
    else:
        raise PermissionDenied


# wallet management payoff page render
@login_required(login_url='user:login') #redirect when user is not logged in
def payoff(request):
    if not request.user.employee.WalletPageAccess :
        raise PermissionDenied
    request.session["activated_menu"]="wallet_manage"
    context={}


    context['wallets']=Wallet.objects.filter(active = True, archived = False, master=True)
    
    context['payoffs']= PayOff.objects.all()

    if 'payoff' in request.GET and PayOff.objects.filter(pk = int(request.GET['payoff'])).exists() :
        selected_payoff = PayOff.objects.get(pk = int(request.GET['payoff']))
        context['selected_payoff'] = selected_payoff
        selected_wallet = selected_payoff.wallet
        context['selected_wallet'] = selected_wallet
        selected_wallet_sum_amount = decimal.Decimal(0)
        selected_wallet_sum_fee = decimal.Decimal(0)

        if selected_wallet.wallet_type == 0 :
            final_tranactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,payed_off = True , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                    .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))
            for fin_trans in final_tranactions:
                temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                fin_trans.end_date = ConvertToSolarDate(temp_start)
                
                if temp_start.month > 1 :
                    temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                else:
                    temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                fin_trans.start_date = ConvertToSolarDate(temp_start)
                selected_wallet_sum_amount += fin_trans.amount
                selected_wallet_sum_fee += fin_trans.fee

            context['final_tranactions'] = final_tranactions                
            context["selected_wallet_sum_amount"]=selected_wallet_sum_amount
            context["selected_wallet_sum_fee"]=selected_wallet_sum_fee
            context["selected_wallet_sum_total"]=selected_wallet_sum_fee + selected_wallet_sum_amount
        else:
            if selected_payoff.balance_after == 0:
                payedoff_transactions = Transaction.objects.filter(incordec = 1,payed_off = True, payoff_wallet = selected_wallet)

                first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                for t in first_transactions:
                    selected_wallet_sum_amount += t.amount
                    selected_wallet_sum_fee += t.fee
                context['first_transactions'] = first_transactions
                context["selected_wallet_sum_amount"]=selected_wallet_sum_amount
                context["selected_wallet_sum_fee"]=selected_wallet_sum_fee
                context["selected_wallet_sum_total"]=selected_wallet_sum_fee + selected_wallet_sum_amount

                prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(next_wallet_id = Value(0,IntegerField()))\
                        .annotate(balance_payoff = Value(0,DecimalField()))

                second_transactions = Transaction.objects.filter(incordec = 1, wallet__in = prime_wallets,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                context['second_transactions'] = second_transactions

                for prime_w in prime_wallets:
                    prime_w.balance_payoff = prime_w.current_balance_payoff(selected_wallet.id)
                    pw_trans = first_transactions.filter(dest_transaction__wallet__id = prime_w.id)
                    for pwt in pw_trans:
                        prime_w.sum_in_amount += pwt.amount
                        prime_w.sum_in_fee += pwt.fee
                    
                    sw_trans = second_transactions.filter(wallet__id = prime_w.id)
                    if sw_trans.count()> 0:
                        prime_w.next_wallet_id = sw_trans.first().dest_transaction.wallet.id

                    current_payedoff = payedoff_transactions.filter(wallet = sw_trans.first().dest_transaction.wallet,pay_off__created__lt = selected_payoff.created)
                    sum_payedoff = 0
                    for cpt in  current_payedoff :
                        sum_payedoff += cpt.amount
                        sum_payedoff += cpt.fee

                    prime_w.sum_in_amount -= sum_payedoff
                    

                context['prime_wallets'] = prime_wallets

                
                final_wallets = Wallet.objects.filter(pk__in = second_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(balance_payoff = Value(0,DecimalField()))

                

                for final_w in final_wallets:
                    final_w.balance_payoff = final_w.current_balance_payoff(selected_wallet.id)
                    fw_trans = second_transactions.filter(dest_transaction__wallet__id = final_w.id)
                    for fwt in fw_trans:
                        final_w.sum_in_amount += fwt.amount
                        final_w.sum_in_fee += fwt.fee

                    current_payedoff = payedoff_transactions.filter(wallet = final_w,pay_off__created__lt = selected_payoff.created)
                    sum_payedoff = 0
                    for cpt in  current_payedoff :
                        sum_payedoff += cpt.amount
                        sum_payedoff += cpt.fee

                    prime_w.sum_in_amount -= sum_payedoff

                context['final_wallets'] = final_wallets
                
                final_tranactions = Transaction.objects.filter(incordec = 1, wallet__in = final_wallets,pay_off = selected_payoff , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                    .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))
                for fin_trans in final_tranactions:
                    temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                    fin_trans.end_date = ConvertToSolarDate(temp_start)
                    
                    if temp_start.month > 1 :
                        try:
                            temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                        except:
                            try:
                                temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day-1)
                            except:
                                temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day-2)
                    else:
                        temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                    fin_trans.start_date = ConvertToSolarDate(temp_start)
                context['final_tranactions'] = final_tranactions
            else :
                payedoff_transactions = Transaction.objects.filter(incordec = 1,payed_off = True, payoff_wallet = selected_wallet)

                selected_wallet_sum_amount = decimal.Decimal(0)
                selected_wallet_sum_fee = decimal.Decimal(0)
                first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                for t in first_transactions:
                    selected_wallet_sum_amount += t.amount
                    selected_wallet_sum_fee += t.fee
                context['first_transactions'] = first_transactions
                context["selected_wallet_sum_amount"]=selected_wallet_sum_amount
                context["selected_wallet_sum_fee"]=selected_wallet_sum_fee
                context["selected_wallet_sum_total"]=selected_wallet_sum_fee + selected_wallet_sum_amount

                prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(next_wallet_id = Value(0,IntegerField()))\
                        .annotate(balance_payoff = Value(0,DecimalField()))

                second_transactions = Transaction.objects.filter(incordec = 1, wallet__in = prime_wallets,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                context['second_transactions'] = second_transactions

                for prime_w in prime_wallets:
                    prime_w.balance_payoff = prime_w.current_balance_payoff(selected_wallet.id)
                    pw_trans = first_transactions.filter(dest_transaction__wallet__id = prime_w.id)
                    for pwt in pw_trans:
                        prime_w.sum_in_amount += pwt.amount
                        prime_w.sum_in_fee += pwt.fee
                    
                    sw_trans = second_transactions.filter(wallet__id = prime_w.id)
                    if sw_trans.count()> 0:
                        prime_w.next_wallet_id = sw_trans.first().dest_transaction.wallet.id

                    current_payedoff = payedoff_transactions.filter(wallet = sw_trans.first().dest_transaction.wallet)
                    sum_payedoff = 0
                    for cpt in  current_payedoff :
                        sum_payedoff += cpt.amount
                        sum_payedoff += cpt.fee

                    prime_w.sum_in_amount -= sum_payedoff
                    

                context['prime_wallets'] = prime_wallets

                
                final_wallets = Wallet.objects.filter(pk__in = second_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(balance_payoff = Value(0,DecimalField()))

                

                for final_w in final_wallets:
                    final_w.balance_payoff = final_w.current_balance_payoff(selected_wallet.id)
                    fw_trans = second_transactions.filter(dest_transaction__wallet__id = final_w.id)
                    for fwt in fw_trans:
                        final_w.sum_in_amount += fwt.amount
                        final_w.sum_in_fee += fwt.fee

                    current_payedoff = payedoff_transactions.filter(wallet = final_w)
                    sum_payedoff = 0
                    for cpt in  current_payedoff :
                        sum_payedoff += cpt.amount
                        sum_payedoff += cpt.fee

                    prime_w.sum_in_amount -= sum_payedoff

                context['final_wallets'] = final_wallets
                
                final_tranactions = Transaction.objects.filter(incordec = 1, wallet__in = final_wallets,payed_off = False , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                    .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))
                for fin_trans in final_tranactions:
                    temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                    fin_trans.end_date = ConvertToSolarDate(temp_start)
                    
                    if temp_start.month > 1 :
                        temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                    else:
                        temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                    fin_trans.start_date = ConvertToSolarDate(temp_start)
                context['final_tranactions'] = final_tranactions
    else:
        try:
            if 'wallet' in request.GET and Wallet.objects.filter(pk = int(request.GET['wallet']) , active = True, archived = False, master=True).exists() :
                selected_wallet= Wallet.objects.get(pk = int(request.GET['wallet']) , active = True, archived = False, master=True)

            else :
                selected_wallet = Wallet.objects.filter(active = True, archived = False, master=True).order_by('current_balance_d')[0]

            context['selected_wallet'] = selected_wallet

            payedoff_transactions = Transaction.objects.filter(incordec = 1,payed_off = True, payoff_wallet = selected_wallet)

            selected_wallet_sum_amount = decimal.Decimal(0)
            selected_wallet_sum_fee = decimal.Decimal(0)

            if selected_wallet.wallet_type == 0 :
                final_tranactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,payed_off = False , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                    .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))
                for fin_trans in final_tranactions:
                    temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                    fin_trans.end_date = ConvertToSolarDate(temp_start)
                    
                    if temp_start.month > 1 :
                        temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                    else:
                        temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                    fin_trans.start_date = ConvertToSolarDate(temp_start)
                    selected_wallet_sum_amount += fin_trans.amount
                    selected_wallet_sum_fee += fin_trans.fee

                context['final_tranactions'] = final_tranactions                
                context["selected_wallet_sum_amount"]=selected_wallet_sum_amount
                context["selected_wallet_sum_fee"]=selected_wallet_sum_fee
                context["selected_wallet_sum_total"]=selected_wallet_sum_fee + selected_wallet_sum_amount

            else:
                first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,payed_off = False, payoff_wallet = selected_wallet)
                for t in first_transactions:
                    selected_wallet_sum_amount += t.amount
                    selected_wallet_sum_fee += t.fee
                context['first_transactions'] = first_transactions
                context["selected_wallet_sum_amount"]=selected_wallet_sum_amount
                context["selected_wallet_sum_fee"]=selected_wallet_sum_fee
                context["selected_wallet_sum_total"]=selected_wallet_sum_fee + selected_wallet_sum_amount

                prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(next_wallet_id = Value(0,IntegerField()))\
                        .annotate(balance_payoff = Value(0,DecimalField()))

                second_transactions = Transaction.objects.filter(incordec = 1, wallet__in = prime_wallets,payed_off = False, payoff_wallet = selected_wallet)
                context['second_transactions'] = second_transactions

                for prime_w in prime_wallets:
                    prime_w.balance_payoff = prime_w.current_balance_payoff(selected_wallet.id)
                    pw_trans = first_transactions.filter(dest_transaction__wallet__id = prime_w.id)
                    for pwt in pw_trans:
                        prime_w.sum_in_amount += pwt.amount
                        prime_w.sum_in_fee += pwt.fee
                    
                    sw_trans = second_transactions.filter(wallet__id = prime_w.id)
                    if sw_trans.count()> 0:
                        prime_w.next_wallet_id = sw_trans.first().dest_transaction.wallet.id

                        current_payedoff = payedoff_transactions.filter(wallet = sw_trans.first().dest_transaction.wallet)
                        sum_payedoff = 0
                        for cpt in  current_payedoff :
                            sum_payedoff += cpt.amount
                            sum_payedoff += cpt.fee

                        prime_w.sum_in_amount -= sum_payedoff
                    

                context['prime_wallets'] = prime_wallets

                
                final_wallets = Wallet.objects.filter(pk__in = second_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
                    .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField())).annotate(balance_payoff = Value(0,DecimalField()))

                

                for final_w in final_wallets:
                    final_w.balance_payoff = final_w.current_balance_payoff(selected_wallet.id)
                    fw_trans = second_transactions.filter(dest_transaction__wallet__id = final_w.id)
                    for fwt in fw_trans:
                        final_w.sum_in_amount += fwt.amount
                        final_w.sum_in_fee += fwt.fee

                    current_payedoff = payedoff_transactions.filter(wallet = final_w)
                    sum_payedoff = 0
                    for cpt in  current_payedoff :
                        sum_payedoff += cpt.amount
                        sum_payedoff += cpt.fee

                    prime_w.sum_in_amount -= sum_payedoff

                context['final_wallets'] = final_wallets
                
                final_tranactions = Transaction.objects.filter(incordec = 1, wallet__in = final_wallets,payed_off = False , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                    .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))
                for fin_trans in final_tranactions:
                    temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                    fin_trans.end_date = ConvertToSolarDate(temp_start)
                    
                    if temp_start.month > 1 :
                        try:
                            temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                        except:
                            temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day - 2 )
                    else:
                        temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                    fin_trans.start_date = ConvertToSolarDate(temp_start)
                context['final_tranactions'] = final_tranactions

        except:
            pass


 

    return render(request, 'wallet/payoff.html', {'context':context})

# wallet management add new additional receipt image to transaction
@login_required(login_url='user:login') #redirect when user is not logged in
def add_receipt(request):
    if request.method == 'POST' and request.user.pk in Permission.objects.filter(codename = 'add_transaction')[0].user_set.all().values_list('pk', flat = True):
        if 'additional_receipt_transaction' in request.POST and request.POST['additional_receipt_transaction']:
            if 'additional_receipt_file' in request.FILES and len(request.FILES.getlist('additional_receipt_file')) > 0:
        
                transaction = Transaction.objects.get(pk = int(request.POST['additional_receipt_transaction']))
                for add_file in request.FILES.getlist('additional_receipt_file'):
                    additional_file = TransactionAdditionalReceipt()
                    additional_file.receipt_file = add_file
                    additional_file.receipt_file_name = add_file.name
                    additional_file.transaction = transaction
                    additional_file.creator = request.user
                    additional_file.save()
        else:
            raise PermissionDenied()

        return redirect("/wallet")
    else:
        raise PermissionDenied()

@login_required(login_url='user:login') #redirect when user is not logged in
def remove_receipt(request, receipt_id):
    try:
        receipt = TransactionAdditionalReceipt.objects.get(pk = receipt_id)
        if receipt.transaction.creator == request.user :
            
            receipt.delete()
            return  HttpResponse(True)
        else:
            raise PermissionDenied
    except Exception as err:
        return  HttpResponse(False)
            
    return  HttpResponse(False)

@login_required(login_url='user:login') #redirect when user is not logged in
def payoff_save(request, wallet_id, payoff_final):
    data = {}
    if request.method == 'POST' and request.user.pk in Permission.objects.filter(codename = 'add_transaction')[0].user_set.all().values_list('pk', flat = True):
        try:
            selected_wallet = Wallet.objects.get(active = True, archived = False, master=True, pk = wallet_id)
        except:
            data['status']="fail"
            data['message']='شماسه کیف پول معتبر نیست'
            return JsonResponse(data)

        if selected_wallet.wallet_type == 0:
            transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,payed_off = False, payoff_wallet = selected_wallet)
            if payoff_final and selected_wallet.current_balance > 0 :
                data['status']="fail"
                data['message']='موجودی کیف پول مستر باید صفر باشد'
                return JsonResponse(data)

            sum_payoff = 0
            for ft in transactions:
                sum_payoff += ft.amount
                sum_payoff += ft.fee

            payoff = PayOff()
            
            _count = PayOff.objects.filter(wallet = selected_wallet).count()

            if payoff_final :
                payoff.title = "تسویه نهایی " + selected_wallet.name
            else:
                payoff.title = "تسویه شماره " + str(_count+1) +" "+ selected_wallet.name

            payoff.wallet = selected_wallet
            payoff.date = datetime.date.today()
            payoff.amount = sum_payoff
            if _count > 0:
                payoff.balance_after = PayOff.objects.filter(wallet = selected_wallet).order_by("-date")[0].balance_after - sum_payoff
            else:
                payoff.balance_after = selected_wallet.initial_balance - sum_payoff

            payoff.creator = request.user

            if 'payoff_title' in request.POST and len(request.POST['payoff_title']) > 0 :
                payoff.title = request.POST['payoff_title']
            if 'payoff_comment' in request.POST and len(request.POST['payoff_comment']) > 0 :
                payoff.comment = request.POST['payoff_comment']

            payoff.save()

            for ft in transactions:
                ft.payed_off = True
                ft.pay_off = payoff
                ft.save()

            if payoff_final :
                selected_wallet.archived = True
                selected_wallet.save()
        else:
            first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,payed_off = False, payoff_wallet = selected_wallet)

            prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))

            if payoff_final :
                if selected_wallet.current_balance > 0 :
                    data['status']="fail"
                    data['message']='موجودی کیف پول مستر باید صفر باشد'
                    return JsonResponse(data)

                for prime_w in prime_wallets:
                    if prime_w.current_balance_payoff(selected_wallet.id) > 0 :
                        data['status']="fail"
                        data['message']='موجودی کیف پول های اولیه باید صفر باشد'
                        return JsonResponse(data)
            else:
                prime_id = []
                for prime_w in prime_wallets:
                    if prime_w.current_balance == 0:
                        prime_id.append(prime_w.id)
                if len(prime_id) == 0:
                    data['status']="fail"
                    data['message']='حداقل موجودی یک کیف پول اولیه باید صفر باشد'
                    return JsonResponse(data)
                prime_wallets = prime_wallets.filter(id__in = prime_id)
                first_transactions = first_transactions.filter(dest_transaction__wallet_id__in = prime_wallets)
            
            second_transactions = Transaction.objects.filter(incordec = 1, wallet__in = prime_wallets,payed_off = False, payoff_wallet = selected_wallet)

            
            final_wallets = Wallet.objects.filter(pk__in = second_transactions.values_list('dest_transaction__wallet__pk',flat=True))


            final_tranactions = Transaction.objects.filter(incordec = 1, wallet__in = final_wallets,payed_off = False , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
            
            sum_payoff = 0
            for ft in final_tranactions:
                sum_payoff += ft.amount
                sum_payoff += ft.fee

            
            for ft in first_transactions:
                sum_payoff += ft.fee

            for st in second_transactions:
                sum_payoff += st.fee

            for fw in final_wallets:
                sum_payoff += fw.current_balance

            payoff = PayOff()
            
            _count = PayOff.objects.filter(wallet = selected_wallet).count()

            if payoff_final :
                payoff.title = "تسویه نهایی " + selected_wallet.name
            else:
                payoff.title = "تسویه شماره " + str(_count+1) +" "+ selected_wallet.name

            payoff.wallet = selected_wallet
            payoff.date = datetime.date.today()
            payoff.amount = sum_payoff
            if _count > 0:
                payoff.balance_after = PayOff.objects.filter(wallet = selected_wallet).order_by("-date")[0].balance_after - sum_payoff
            else:
                payoff.balance_after = selected_wallet.initial_balance - sum_payoff

            payoff.creator = request.user

            if 'payoff_title' in request.POST and len(request.POST['payoff_title']) > 0 :
                payoff.title = request.POST['payoff_title']
            if 'payoff_comment' in request.POST and len(request.POST['payoff_comment']) > 0 :
                payoff.comment = request.POST['payoff_comment']

            payoff.save()

            for ft in final_tranactions:
                ft.payed_off = True
                ft.pay_off = payoff
                ft.save()

            
            for ft in first_transactions:
                ft.payed_off = True
                ft.pay_off = payoff
                ft.save()
                ft.dest_transaction.payed_off = True
                ft.dest_transaction.pay_off = payoff
                ft.dest_transaction.save()

            for st in second_transactions:
                st.payed_off = True
                st.pay_off = payoff
                st.save() 
                st.dest_transaction.payed_off = True
                st.dest_transaction.pay_off = payoff
                st.dest_transaction.save()   

            for fw in final_wallets :
                fw.archived = True
                fw.save()

            if payoff_final:
                selected_wallet.archived = True
                selected_wallet.save()
            
        data['status']="success"
        data['message']='تسویه با موفقیت ثبت شد'
        return JsonResponse(data)
    else:
        raise PermissionDenied()


@login_required(login_url='user:login') #redirect when user is not logged in
def download_payoff_excell(request,payoff_id):
    selected_payoff = PayOff.objects.get(pk = payoff_id)
    selected_wallet = selected_payoff.wallet
    payedoff_transactions = Transaction.objects.filter(incordec = 1,payed_off = True, payoff_wallet = selected_wallet)

    center_align = alignment=Alignment(horizontal='center')
    right_align = alignment=Alignment(horizontal='right')



    wb = Workbook()
    wb.create_sheet
    sheet_count = 0
    
    selected_wallet_sum_amount = decimal.Decimal(0)
    selected_wallet_sum_fee = decimal.Decimal(0)
    if selected_wallet.wallet_type == 0:
        sheet_count += 1
        if sheet_count == 1:
            sheet = wb.active
            sheet.title = "1"
        else :
            sheet = wb.create_sheet(title = str(sheet_count))

        sheet.sheet_view.rightToLeft = True
        total_fee = 0
        selected_wallet.balance_payoff = selected_wallet.current_balance_payoff(selected_wallet.id)


        final_tranactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
            .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))

        a1 = sheet['A1']
        b1 = sheet['B1']
        c1 = sheet['C1']
        d1 = sheet['D1']
        e1 = sheet['E1']

        a1.value = "ردیف"
        b1.value = "کشور"
        c1.value = "تاریخ شروع"
        d1.value = "تاریخ پایان"
        e1.value = "هزینه سرور"

        a1.alignment = center_align
        a1.font = Font(name='B Titr', size=13, bold=True)
        a1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        b1.alignment = center_align
        b1.font = Font(name='B Titr', size=13, bold=True)
        b1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        c1.alignment = center_align
        c1.font = Font(name='B Titr', size=13, bold=True)
        c1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        d1.alignment = center_align
        d1.font = Font(name='B Titr', size=13, bold=True)
        d1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        e1.alignment = center_align
        e1.font = Font(name='B Titr', size=13, bold=True)
        e1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20

        row_count = 0
        total_pay = 0

        for fin_trans in final_tranactions:
            row_count += 1
            an = sheet['A'+str(row_count+1)]
            bn = sheet['B'+str(row_count+1)]
            cn = sheet['C'+str(row_count+1)]
            dn = sheet['D'+str(row_count+1)]
            en = sheet['E'+str(row_count+1)]


            total_fee += fin_trans.fee
            total_pay += fin_trans.amount

            if fin_trans.dest_resource.payoff_code:
                payoff_code = fin_trans.dest_resource.payoff_code
            else:
                max_code = Resource.objects.all().exclude(payoff_code = None).order_by('-payoff_code').first().payoff_code
                fin_trans.dest_resource.payoff_code = max_code + 1
                fin_trans.dest_resource.save()
                payoff_code = max_code + 1

            temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
            fin_trans.end_date = ConvertToSolarDate(temp_start)
            
            if temp_start.month > 1 :
                temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
            else:
                temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
            fin_trans.start_date = ConvertToSolarDate(temp_start)

            try:
                ram = int(fin_trans.dest_resource.resource_num_properties.all().filter(resource_type_property__slug = "vps_ram").first().value)
            except:
                ram = 0

            try:
                storage = int(fin_trans.dest_resource.resource_num_properties.all().filter(resource_type_property__slug = "vps_storage").first().value)
            except:
                storage = 0

            try:
                os_n = fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "vps-OS").first().value
                if os_n.lower().find('win') > -1:
                    os_n = "W"
                else:
                    os_n = "L"
            except:
                if fin_trans.dest_resource.resource_type.slug == 'Tell':
                    os_n = 'PH'
                else:
                    os_n = 'D'
                    temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                    temp_start = datetime.date(year= temp_start.year - 1, month=temp_start.month , day=temp_start.day)
                    fin_trans.start_date = ConvertToSolarDate(temp_start)

            an.value = row_count
            an.font = Font(name='B Titr', size=12, bold=True)
            an.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            bn.value = "02/" + os_n + "/" + str(ram) + "/0/" + str(storage) + "/ " + str(fin_trans.dest_resource.payoff_code)
            bn.alignment = center_align
            bn.font = Font(name='B Kamran', size=12, bold=True)

            cn.value = fin_trans.start_date
            cn.alignment = center_align
            cn.font = Font(name='B Kamran', size=12, bold=True)

            dn.value = fin_trans.end_date 
            dn.alignment = center_align
            dn.font = Font(name='B Kamran', size=12, bold=True)

            en.value = fin_trans.amount
            en.alignment = center_align
            en.font = Font(name='B Titr', size=12, bold=True)
    
        an1 = sheet['A'+str(row_count+2)]
        an2 = sheet['A'+str(row_count+3)]
        an3 = sheet['A'+str(row_count+4)]
        ant = sheet['A'+str(row_count+5)]
        bnt = sheet['B'+str(row_count+5)]
        ent = sheet['E'+str(row_count+5)]

        ans = sheet['A'+str(row_count+6)]
        ens = sheet['E'+str(row_count+6)]

        anw = sheet['A'+str(row_count+7)]
        anx = sheet['A'+str(row_count+8)]
        bnx = sheet['B'+str(row_count+8)]
        cnx = sheet['C'+str(row_count+8)]
        dnx = sheet['D'+str(row_count+8)]
        _any = sheet['A'+str(row_count+9)]
        bny = sheet['B'+str(row_count+9)]
        cny = sheet['C'+str(row_count+9)]
        dny = sheet['D'+str(row_count+9)]
        anz = sheet['A'+str(row_count+10)]
        dnz = sheet['D'+str(row_count+10)]


        sheet.merge_cells('A'+str(row_count+6)+':'+'D'+str(row_count+6))
        sheet.merge_cells('A'+str(row_count+7)+':'+'E'+str(row_count+7))
        sheet.merge_cells('D'+str(row_count+8)+':'+'E'+str(row_count+8))
        sheet.merge_cells('D'+str(row_count+9)+':'+'E'+str(row_count+9))
        sheet.merge_cells('A'+str(row_count+10)+':'+'C'+str(row_count+10))
        sheet.merge_cells('D'+str(row_count+10)+':'+'E'+str(row_count+10))

        an1.value = row_count + 1
        an1.font = Font(name='B Titr', size=12, bold=True)
        an1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        an2.value = row_count + 2
        an2.font = Font(name='B Titr', size=12, bold=True)
        an2.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        an3.value = row_count + 3
        an3.font = Font(name='B Titr', size=12, bold=True)
        an3.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        ant.value = row_count + 4 
        ant.font = Font(name='B Titr', size=12, bold=True)
        ant.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

        bnt.value = "Total Fee"
        bnt.font = Font(name='B Titr', size=12, bold=True)
        bnt.alignment = center_align

        ent.value = total_fee
        ent.font = Font(name='B Titr', size=12, bold=True)
        ent.alignment = center_align

        ans.value = "جمع کل"
        ans.font = Font(name='B Titr', size=12, bold=True)
        ans.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        ans.alignment = center_align

        ens.value = total_pay + total_fee
        ens.font = Font(name='B Titr', size=12, bold=True)
        ens.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        ens.alignment = center_align

        anw.value = "هزینه های صورت گرفته در فوق توسط کارت های زیر صورت پذیرفته است:"
        anw.font = Font(name='B Titr', size=13, bold=True)

        anx.value = "ردیف"
        anx.font = Font(name='B Titr', size=11, bold=True)
        anx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        anx.alignment = center_align

        bnx.value = "شماره کارت"
        bnx.font = Font(name='B Titr', size=11, bold=True)
        bnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        bnx.alignment = center_align

        cnx.value = "مبلغ اولیه کارت"
        cnx.font = Font(name='B Titr', size=11, bold=True)
        cnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        cnx.alignment = center_align

        dnx.value = "مبلغ استفاده شده برای سرور های فوق"
        dnx.font = Font(name='B Titr', size=11, bold=True)
        dnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        dnx.alignment = center_align

        _any.value = "1"
        _any.font = Font(name='B Titr', size=11, bold=True)
        _any.alignment = center_align

        bny.value = selected_wallet.name
        bny.font = Font(name='B Titr', size=11, bold=True)
        bny.alignment = center_align

        cny.value = selected_wallet.initial_balance
        cny.font = Font(name='B Titr', size=11, bold=True)
        cny.alignment = center_align

        dny.value = total_pay + total_fee
        dny.font = Font(name='B Titr', size=12, bold=True)
        dny.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        dny.alignment = center_align

        anz.value = "جمع کل"
        anz.font = Font(name='B Titr', size=11, bold=True)
        anz.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        anz.alignment = center_align

        dnz.value = total_pay + total_fee
        dnz.font = Font(name='B Titr', size=12, bold=True)
        dnz.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
        dnz.alignment = center_align
    else:
        
        first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff, payoff_wallet = selected_wallet)        

        prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))\
            .annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField()))
        
        if selected_payoff.balance_after != 0:
            prime_id = []
            for prime_w in prime_wallets:
                if prime_w.current_balance == 0:
                    prime_id.append(prime_w.id)
            
            prime_wallets = prime_wallets.filter(id__in = prime_id)
            first_transactions = first_transactions.filter(dest_transaction__wallet_id__in = prime_wallets)

        for prime_w in prime_wallets:
            sheet_count += 1
            if sheet_count == 1:
                sheet = wb.active
                sheet.title = "1"
            else :
                sheet = wb.create_sheet(title = str(sheet_count))

            sheet.sheet_view.rightToLeft = True
            total_fee = 0
            prime_w.balance_payoff = prime_w.current_balance_payoff(selected_wallet.id)
            pw_trans = first_transactions.filter(dest_transaction__wallet__id = prime_w.id)
            for pwt in pw_trans:
                total_fee += pwt.fee

            second_transactions = Transaction.objects.filter(incordec = 1, wallet = prime_w,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                        
            final_wallet = second_transactions.first().dest_transaction.wallet
            total_fee += final_wallet.current_balance
            fw_trans = second_transactions.filter(dest_transaction__wallet__id = final_wallet.id)
            for fwt in fw_trans:
                total_fee += fwt.fee

            final_tranactions = Transaction.objects.filter(incordec = 1, wallet = final_wallet,pay_off = selected_payoff , payoff_wallet = selected_wallet).exclude(dest_resource = None)\
                .annotate(start_date= Value("",CharField())).annotate(end_date= Value("",CharField()))

            a1 = sheet['A1']
            b1 = sheet['B1']
            c1 = sheet['C1']
            d1 = sheet['D1']
            e1 = sheet['E1']

            a1.value = "ردیف"
            b1.value = "کشور"
            c1.value = "تاریخ شروع"
            d1.value = "تاریخ پایان"
            e1.value = "هزینه سرور"

            a1.alignment = center_align
            a1.font = Font(name='B Titr', size=13, bold=True)
            a1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            b1.alignment = center_align
            b1.font = Font(name='B Titr', size=13, bold=True)
            b1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            c1.alignment = center_align
            c1.font = Font(name='B Titr', size=13, bold=True)
            c1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            d1.alignment = center_align
            d1.font = Font(name='B Titr', size=13, bold=True)
            d1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            e1.alignment = center_align
            e1.font = Font(name='B Titr', size=13, bold=True)
            e1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            sheet.column_dimensions["A"].width = 5
            sheet.column_dimensions["B"].width = 40
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 20
            sheet.column_dimensions["E"].width = 20

            row_count = 0
            total_pay = 0

            for fin_trans in final_tranactions:
                row_count += 1
                an = sheet['A'+str(row_count+1)]
                bn = sheet['B'+str(row_count+1)]
                cn = sheet['C'+str(row_count+1)]
                dn = sheet['D'+str(row_count+1)]
                en = sheet['E'+str(row_count+1)]


                total_fee += fin_trans.fee
                total_pay += fin_trans.amount

                if fin_trans.dest_resource.payoff_code:
                    payoff_code = fin_trans.dest_resource.payoff_code
                else:
                    max_code = Resource.objects.all().exclude(payoff_code = None).order_by('-payoff_code').first().payoff_code
                    fin_trans.dest_resource.payoff_code = max_code + 1
                    fin_trans.dest_resource.save()
                    payoff_code = max_code + 1

                temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                fin_trans.end_date = ConvertToSolarDate(temp_start)
                
                if temp_start.month > 1 :
                    try:
                        temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day)
                    except:
                        try:
                            temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day - 1)
                        except:
                            temp_start = datetime.date(year= temp_start.year, month=temp_start.month -1 , day=temp_start.day - 2)
                else:
                    temp_start = datetime.date(year= temp_start.year -1 , month=12 , day=temp_start.day)
                fin_trans.start_date = ConvertToSolarDate(temp_start)

                try:
                    ram = int(fin_trans.dest_resource.resource_num_properties.all().filter(resource_type_property__slug = "vps_ram").first().value)
                except:
                    ram = 0

                try:
                    storage = int(fin_trans.dest_resource.resource_num_properties.all().filter(resource_type_property__slug = "vps_storage").first().value)
                except:
                    storage = 0

                try:
                    os_n = fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "vps-OS").first().value
                    if os_n.lower().find('win') > -1:
                        os_n = "W"
                    else:
                        os_n = "L"
                except:
                    if fin_trans.dest_resource.resource_type.slug == 'Tell':
                        os_n = 'PH'
                    else:
                        os_n = 'D'
                        temp_start = fin_trans.dest_resource.consuming_resources.all().order_by("-expiration").first().expiration
                        temp_start = datetime.date(year= temp_start.year - 1, month=temp_start.month , day=temp_start.day)
                        fin_trans.start_date = ConvertToSolarDate(temp_start)

                an.value = row_count
                an.font = Font(name='B Titr', size=12, bold=True)
                an.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

                bn.value = "02/" + os_n + "/" + str(ram) + "/0/" + str(storage) + "/ " + str(fin_trans.dest_resource.payoff_code)
                bn.alignment = center_align
                bn.font = Font(name='B Kamran', size=12, bold=True)

                cn.value = fin_trans.start_date
                cn.alignment = center_align
                cn.font = Font(name='B Kamran', size=12, bold=True)

                dn.value = fin_trans.end_date 
                dn.alignment = center_align
                dn.font = Font(name='B Kamran', size=12, bold=True)

                en.value = fin_trans.amount
                en.alignment = center_align
                en.font = Font(name='B Titr', size=12, bold=True)
        
            an1 = sheet['A'+str(row_count+2)]
            an2 = sheet['A'+str(row_count+3)]
            an3 = sheet['A'+str(row_count+4)]
            ant = sheet['A'+str(row_count+5)]
            bnt = sheet['B'+str(row_count+5)]
            ent = sheet['E'+str(row_count+5)]

            ans = sheet['A'+str(row_count+6)]
            ens = sheet['E'+str(row_count+6)]

            anw = sheet['A'+str(row_count+7)]
            anx = sheet['A'+str(row_count+8)]
            bnx = sheet['B'+str(row_count+8)]
            cnx = sheet['C'+str(row_count+8)]
            dnx = sheet['D'+str(row_count+8)]
            _any = sheet['A'+str(row_count+9)]
            bny = sheet['B'+str(row_count+9)]
            cny = sheet['C'+str(row_count+9)]
            dny = sheet['D'+str(row_count+9)]
            anz = sheet['A'+str(row_count+10)]
            dnz = sheet['D'+str(row_count+10)]


            sheet.merge_cells('A'+str(row_count+6)+':'+'D'+str(row_count+6))
            sheet.merge_cells('A'+str(row_count+7)+':'+'E'+str(row_count+7))
            sheet.merge_cells('D'+str(row_count+8)+':'+'E'+str(row_count+8))
            sheet.merge_cells('D'+str(row_count+9)+':'+'E'+str(row_count+9))
            sheet.merge_cells('A'+str(row_count+10)+':'+'C'+str(row_count+10))
            sheet.merge_cells('D'+str(row_count+10)+':'+'E'+str(row_count+10))

            an1.value = row_count + 1
            an1.font = Font(name='B Titr', size=12, bold=True)
            an1.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            an2.value = row_count + 2
            an2.font = Font(name='B Titr', size=12, bold=True)
            an2.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            an3.value = row_count + 3
            an3.font = Font(name='B Titr', size=12, bold=True)
            an3.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            ant.value = row_count + 4 
            ant.font = Font(name='B Titr', size=12, bold=True)
            ant.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')

            bnt.value = "Total Fee"
            bnt.font = Font(name='B Titr', size=12, bold=True)
            bnt.alignment = center_align

            ent.value = total_fee
            ent.font = Font(name='B Titr', size=12, bold=True)
            ent.alignment = center_align

            ans.value = "جمع کل"
            ans.font = Font(name='B Titr', size=12, bold=True)
            ans.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            ans.alignment = center_align

            ens.value = total_pay + total_fee
            ens.font = Font(name='B Titr', size=12, bold=True)
            ens.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            ens.alignment = center_align

            anw.value = "هزینه های صورت گرفته در فوق توسط کارت های زیر صورت پذیرفته است:"
            anw.font = Font(name='B Titr', size=13, bold=True)

            anx.value = "ردیف"
            anx.font = Font(name='B Titr', size=11, bold=True)
            anx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            anx.alignment = center_align

            bnx.value = "شماره کارت"
            bnx.font = Font(name='B Titr', size=11, bold=True)
            bnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            bnx.alignment = center_align

            cnx.value = "مبلغ اولیه کارت"
            cnx.font = Font(name='B Titr', size=11, bold=True)
            cnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            cnx.alignment = center_align

            dnx.value = "مبلغ استفاده شده برای سرور های فوق"
            dnx.font = Font(name='B Titr', size=11, bold=True)
            dnx.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            dnx.alignment = center_align

            _any.value = "1"
            _any.font = Font(name='B Titr', size=11, bold=True)
            _any.alignment = center_align

            bny.value = selected_wallet.name
            bny.font = Font(name='B Titr', size=11, bold=True)
            bny.alignment = center_align

            cny.value = selected_wallet.initial_balance
            cny.font = Font(name='B Titr', size=11, bold=True)
            cny.alignment = center_align

            dny.value = total_pay + total_fee
            dny.font = Font(name='B Titr', size=12, bold=True)
            dny.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            dny.alignment = center_align

            anz.value = "جمع کل"
            anz.font = Font(name='B Titr', size=11, bold=True)
            anz.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            anz.alignment = center_align

            dnz.value = total_pay + total_fee
            dnz.font = Font(name='B Titr', size=12, bold=True)
            dnz.fill = PatternFill(fill_type=None, start_color='FFBBBBBB', end_color='FFBBBBBB')
            dnz.alignment = center_align
    


    response = HttpResponse( content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" , )
    response["Content-Disposition"] = "attachment;filename=payoff-{date}.xlsx".format(date = selected_payoff.date.strftime("%Y-%m-%d"),)

    wb.save(response)

    return  response

@login_required(login_url='user:login') #redirect when user is not logged in
def download_payoff_zip(request, payoff_id):
    selected_payoff = PayOff.objects.get(pk = payoff_id)
    selected_wallet = selected_payoff.wallet
    payedoff_transactions = Transaction.objects.filter(incordec = 1,payed_off = True, payoff_wallet = selected_wallet)

    # Folder name in ZIP archive which contains the above files
    # E.g [thearchive.zip]/somefiles/file2.txt
    # FIXME: Set this to something better
    zip_subdir = selected_payoff.title
    zip_filename = "payoff-{date}.zip".format(date = selected_payoff.date.strftime("%Y-%m-%d"),)

    # Open StringIO to grab in-memory ZIP contents
    bio = BytesIO()

    # The zip compressor
    zf = ZipFile(bio, "w")

    selected_wallet_sum_amount = decimal.Decimal(0)
    selected_wallet_sum_fee = decimal.Decimal(0)
    if selected_wallet.wallet_type == 0:
        fin_count = 0
        fin_letter = string.ascii_letters[selected_payoff.pk % 26]
        sub_path = selected_wallet.name.replace('کیف','').replace('پول','').replace('اولیه','').replace('نهایی','').replace(' ','')
        total_fee = 0

        final_tranactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff , payoff_wallet = selected_wallet).exclude(dest_resource = None)

        for fin_trans in final_tranactions:
            fin_count +=1
            total_fee += fin_trans.fee
            try:
                finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "vps-site").first().value
            except:
                try:
                    finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "Ph-site").first().value
                except:
                    try:
                        finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "D-site").first().value
                    except:
                        finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'
            rec_path = fin_trans.receipt_file.file.name
            if fin_trans.receipt_file_name :
                rec_name = fin_trans.receipt_file_name
            else:
                rec_name = 'برداشت' + str(fin_trans.amount.normalize())+'.jpg'
            zf.write(rec_path, os.path.join(zip_subdir,sub_path,finpath,rec_name))
            count = 0
            for add_rec in fin_trans.additional_receipts.all():
                count += 1
                rec_path = add_rec.receipt_file.file.name
                if add_rec.receipt_file_name:
                    rec_name_add = add_rec.receipt_file_name
                else:
                    rec_name_add = rec_name + '_' + str(count)
                zf.write(rec_path, os.path.join(zip_subdir,sub_path,finpath,rec_name_add+'.jpg'))


    else:
        first_transactions = Transaction.objects.filter(incordec = 1, wallet = selected_wallet,pay_off = selected_payoff, payoff_wallet = selected_wallet)        

        prime_wallets = Wallet.objects.filter(pk__in = first_transactions.values_list('dest_transaction__wallet__pk',flat=True))
            
        
        if selected_payoff.balance_after != 0:
            prime_id = []
            for prime_w in prime_wallets:
                if prime_w.current_balance == 0:
                    prime_id.append(prime_w.id)
            
            prime_wallets = prime_wallets.filter(id__in = prime_id)
            first_transactions = first_transactions.filter(dest_transaction__wallet_id__in = prime_wallets)
        
        prime_wallets = prime_wallets.annotate(sum_in_amount = Value(0,IntegerField())).annotate(sum_in_fee = Value(0,IntegerField()))
        fin_count = 0
        fin_letter = string.ascii_letters[selected_payoff.pk % 26]
        
        for prime_w in prime_wallets:

            sub_path = prime_w.name.replace('کیف','').replace('پول','').replace('اولیه','').replace('نهایی','').replace(' ','')
            
            # # Calculate path for file in zip
            # fdir, fname = os.path.split(fpath)
            # zip_path = os.path.join(zip_subdir, fname)

            # # Add file, at correct path
            # zf.write(fpath, zip_path)

            total_fee = 0
            prime_w.balance_payoff = prime_w.current_balance_payoff(selected_wallet.id)
            pw_trans = first_transactions.filter(dest_transaction__wallet__id = prime_w.id)
            for pwt in pw_trans:
                total_fee += pwt.fee
                rec_path = pwt.receipt_file.file.name
                if pwt.receipt_file_name :
                    rec_name = pwt.receipt_file_name
                else:
                    rec_name = 'برداشت' + str(pwt.amount.normalize())+'.jpg'
                zf.write(rec_path, os.path.join(zip_subdir,sub_path,rec_name))
                count = 0
                for add_rec in pwt.additional_receipts.all():
                    count += 1
                    rec_path = add_rec.receipt_file.file.name
                    if add_rec.receipt_file_name:
                        rec_name_add = add_rec.receipt_file_name
                    else:
                        rec_name_add = rec_name + '_' + str(count)
                    zf.write(rec_path, os.path.join(zip_subdir,sub_path,rec_name_add+'.jpg'))

            second_transactions = Transaction.objects.filter(incordec = 1, wallet = prime_w,pay_off = selected_payoff, payoff_wallet = selected_wallet)
                        
            final_wallet = second_transactions.first().dest_transaction.wallet
            fw_trans = second_transactions.filter(dest_transaction__wallet__id = final_wallet.id)
            for fwt in fw_trans:
                total_fee += fwt.fee
                rec_path = fwt.receipt_file.file.name
                if fwt.receipt_file_name :
                    rec_name = fwt.receipt_file_name
                else:
                    rec_name = 'برداشت' + str(fwt.amount.normalize())+'.jpg'
                zf.write(rec_path, os.path.join(zip_subdir,sub_path,rec_name))
                count = 0
                for add_rec in fwt.additional_receipts.all():
                    count += 1
                    rec_path = add_rec.receipt_file.file.name
                    if add_rec.receipt_file_name:
                        rec_name_add = add_rec.receipt_file_name
                    else:
                        rec_name_add = rec_name + '_' + str(count)
                    zf.write(rec_path, os.path.join(zip_subdir,sub_path,rec_name_add+'.jpg'))

            final_tranactions = Transaction.objects.filter(incordec = 1, wallet = final_wallet,pay_off = selected_payoff , payoff_wallet = selected_wallet).exclude(dest_resource = None)

            for fin_trans in final_tranactions:
                fin_count +=1
                total_fee += fin_trans.fee
                try:
                    finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "vps-site").first().value
                except:
                    try:
                        finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "Ph-site").first().value
                    except:
                        try:
                            finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'+fin_trans.dest_resource.resource_text_properties.all().filter(resource_type_property__slug = "D-site").first().value
                        except:
                            finpath = fin_letter+str(fin_count) + '-----'+str((fin_trans.amount+fin_trans.fee).normalize())+'----'
                rec_path = fin_trans.receipt_file.file.name
                if fin_trans.receipt_file_name :
                    rec_name = fin_trans.receipt_file_name
                else:
                    rec_name = 'برداشت' + str(fin_trans.amount.normalize())+'.jpg'
                zf.write(rec_path, os.path.join(zip_subdir,sub_path,finpath,rec_name))
                count = 0
                for add_rec in fin_trans.additional_receipts.all():
                    count += 1
                    rec_path = add_rec.receipt_file.file.name
                    if add_rec.receipt_file_name:
                        rec_name_add = add_rec.receipt_file_name
                    else:
                        rec_name_add = rec_name + '_' + str(count)
                    zf.write(rec_path, os.path.join(zip_subdir,sub_path,finpath,rec_name_add+'.jpg'))






    # for fpath in filenames:
    #     # Calculate path for file in zip
    #     fdir, fname = os.path.split(fpath)
    #     zip_path = os.path.join(zip_subdir, fname)

    #     # Add file, at correct path
    #     zf.write(fpath, zip_path)

    # Must close zip for all contents to be written
    zf.close()

    # Grab ZIP file from in-memory, make response with correct MIME-type
    resp = HttpResponse(bio.getvalue(), content_type = "application/x-zip-compressed")
    # ..and correct content-disposition
    resp['Content-Disposition'] = 'attachment; filename=%s' % zip_filename

    return resp


# wallet management payoff page render
@login_required(login_url='user:login') #redirect when user is not logged in
def report(request):
    if not request.user.employee.WalletPageAccess :
        raise PermissionDenied
    request.session["activated_menu"]="wallet_manage"
    context={}

    gto_date=datetime.datetime.now().date()

    gfrom_date=(datetime.datetime.now() - datetime.timedelta(35)).date()

    if 'from' in request.GET and 'to' in request.GET :
        from_date = request.GET.get('from')
        to_date = request.GET.get('to')
        gto_date = jdt.strptime(to_date,'%Y-%m-%d').togregorian().date()
        gfrom_date = jdt.strptime(from_date,'%Y-%m-%d').togregorian().date()
        if gto_date < gfrom_date :
            gto_date = gfrom_date

    filtered_transactions = Transaction.objects.filter(time__date__gte = gfrom_date , time__date__lte = gto_date, incordec = 1)
    final_tranactions = filtered_transactions.exclude(dest_resource = None)
    mid_transactions = filtered_transactions.filter(dest_resource = None)

    wallets = filtered_transactions.values_list("wallet__id",flat=True)
    areas = OPArea.objects.all().annotate(dollar_amount = Value(0.0,DecimalField())).annotate(other_dollar_amount = Value(0.0,DecimalField()))
    projects = OPProject.objects.all().annotate(dollar_amount = Value(0.0,DecimalField()))


    max_project = 0.0
    for project in projects :
        project_resources = Resource.objects.filter(id__in = project.consuming_resources.all().values_list('resource_id',flat=True))
        total_fee = final_tranactions.filter(dest_resource__in = project_resources).aggregate(Sum('fee_dollar'))
        total_buy = final_tranactions.filter(dest_resource__in = project_resources).aggregate(Sum('amount_dollar'))
        if total_fee['fee_dollar__sum'] and  total_buy['amount_dollar__sum']:
            project.dollar_amount = total_fee['fee_dollar__sum'] + total_buy['amount_dollar__sum']
            if project.dollar_amount > max_project :
                max_project = project.dollar_amount

    total = 0
    for area in areas:
        total_fee = filtered_transactions.filter(wallet__in = area.wallets.all()).aggregate(Sum('fee_dollar'))
        total_buy = final_tranactions.filter(wallet__in = area.wallets.all()).aggregate(Sum('amount_dollar'))
        if total_fee['fee_dollar__sum'] and  total_buy['amount_dollar__sum']:
            area.dollar_amount = total_fee['fee_dollar__sum'] + total_buy['amount_dollar__sum']
            total += area.dollar_amount

            projects_total = 0
            for project in projects:
                if project.area == area:
                    projects_total += project.dollar_amount
            area.other_dollar_amount = area.dollar_amount - projects_total
            if area.other_dollar_amount > max_project:
                max_project = area.other_dollar_amount

    

    context['fyear']=gfrom_date.year
    context['fmonth']=gfrom_date.month-1
    context['fday'] = gfrom_date.day
    context['tyear'] = gto_date.year
    context['tmonth'] = gto_date.month-1
    context['tday'] = gto_date.day
    context['areas'] = areas
    context['projects'] = projects
    context['max_project'] = max_project
    context["total"] = total


    return render(request, 'wallet/report.html', {'context':context})
